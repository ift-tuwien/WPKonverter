"""Convert WPK mail registration data"""

# -- Import -------------------------------------------------------------------

from argparse import ArgumentParser, Namespace
from csv import DictReader
from logging import basicConfig, getLogger

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from pyparsing import col, lineno, ParseException, ParseResults

from wpkonverter.cli import file_exists
from wpkonverter.parsing import mail, mail_attributes

# -- Functions ----------------------------------------------------------------


def get_arguments() -> Namespace:
    """Parse command line arguments

    Returns:

        An object that contains the given command line arguments

    """

    parser = ArgumentParser(
        description="Extract data from WPK registration mails"
    )

    parser.add_argument(
        "--log",
        choices=("debug", "info", "warning", "error", "critical"),
        default="warning",
        required=False,
        help="minimum log level",
    )

    parser.add_argument(
        "filepath",
        type=file_exists,
        help="WPK mail information in CSV format",
    )

    return parser.parse_args()


def generate_error_message(text: str, error: ParseException) -> str:
    """Generate a human readable parsing error message

    Args:

        text:

            The parsed text that produced an error

        error:

            The error that occurred while parsing ``text``

    Returns:

        An error message that shows the user where the parsing error occurred

    """

    line_number = lineno(error.loc, text)
    column_number = col(error.loc, text)
    lines = text.splitlines()
    quote_symbol = "> "
    error_indent = " " * (len(quote_symbol) + column_number)

    error_message: list[str] = []
    for line in lines[line_number - 5 : line_number]:
        error_message.append(f"{quote_symbol}{line}")
    error_message.append(f"{error_indent}^")
    error_message.append(f"{error_indent}{error}")
    for line in lines[line_number : line_number + 1]:
        error_message.append(f"{quote_symbol}{line}")

    return "\n".join(error_message)


def store_data_workbook(parsed_mails: list[ParseResults]) -> None:
    """Store registration data in Excel file

    Args:

        data:

            The registration data that should be stored in the Excel file

    """

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pre-registration"

    worksheet.append([attribute.capitalize() for attribute in mail_attributes])

    bold = Font(bold=True)
    start_column = get_column_letter(worksheet.min_column)
    end_column = get_column_letter(worksheet.max_column)
    for row in worksheet[f"{start_column}1:{end_column}1"]:
        for cell in row:
            cell.font = bold

    dimension_holder = DimensionHolder(worksheet=worksheet)

    for column in range(worksheet.min_column, worksheet.max_column + 1):
        dimension_holder[get_column_letter(column)] = ColumnDimension(
            worksheet, min=column, max=column, width=40
        )

    worksheet.column_dimensions = dimension_holder

    for parsed_mail in parsed_mails:
        worksheet.append(
            [parsed_mail[attribute] for attribute in mail_attributes]
        )

    filename = "wpk.xlsx"
    workbook.save(filename)
    print(f"Stored data in “{filename}”")


# -- Main ---------------------------------------------------------------------


def main() -> None:
    """Convert WPK mail registration data"""

    arguments = get_arguments()

    basicConfig(
        encoding="utf-8",
        format="{asctime} {levelname:7} {message}",
        level=arguments.log.upper(),
        style="{",
    )

    logger = getLogger(__name__)
    logger.info("CLI Arguments: %s", arguments)

    try:
        with open(arguments.filepath, newline="", encoding="utf8") as csvfile:
            reader = DictReader(csvfile)
            parsed_mails: list[ParseResults] = []
            for mail_number, row in enumerate(reader, start=1):
                text = row["Text"]
                logger.debug("Mail text: %s", text)
                try:
                    parsed_mail = mail.parse_string(text, parse_all=True)
                    parsed_mails.append(parsed_mail)

                    for attribute in mail_attributes:
                        logger.debug(
                            "%s: %s", attribute, parsed_mail[attribute]
                        )
                except ParseException as error:
                    print(f"Unable to parse data in mail {mail_number}:\n")
                    print(generate_error_message(text, error))
                    print()

            store_data_workbook(parsed_mails)
    except UnicodeDecodeError as error:
        print(f"Unable to read file “{arguments.filepath}”: {error}")
